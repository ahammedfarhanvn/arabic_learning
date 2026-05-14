from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseBadRequest
from .models import Customer, Item, Transaction, ItemPurchase
from .forms import CustomerForm, ItemForm
from django.http import JsonResponse
from decimal import Decimal
import csv
import io
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from openpyxl import Workbook, styles
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, styles
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# views for displaying the contents of the pages 
def  home(request):
    return render(request, 'store/home.html')

def add_item(request):
    if request.method == "POST":
        form = ItemForm(request.POST)
        if form.is_valid():
            form.save()  # Save the new item to the database
            return redirect('item_list')  # Redirect to a page that lists all items (you can create this later)
    else:
        form = ItemForm()
    return render(request, 'store/add_item.html', {'form': form})

def item_list(request):
    items = Item.objects.all()  # Retrieve all items from the database
    return render(request, 'store/item_list.html', {'items': items})

def edit_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if request.method == "POST":
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()  # Update the item in the database
            return redirect('item_list')  # Redirect to the item list page
    else:
        form = ItemForm(instance=item)
    return render(request, 'store/edit_item.html', {'form': form, 'item': item})

def delete_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if request.method == "POST":
        item.delete()  # Delete the item from the database
        return redirect('item_list')  # Redirect to the item list page
    return render(request, 'store/delete_item.html', {'item': item})

def add_customer(request):
    if request.method == "POST":
        # Handle form submission for adding a single customer
        customer_number = request.POST.get('customer_number')
        name = request.POST.get('name')
        balance = request.POST.get('balance')
        if name and balance.isdigit() and customer_number:
            Customer.objects.create(customer_number=customer_number, name=name, balance=int(balance))
            return redirect('customer_list')

        # Handle file upload
        uploaded_file = request.FILES.get('customer_file')
        if uploaded_file:
            fs = FileSystemStorage()
            filename = fs.save(uploaded_file.name, uploaded_file)
            uploaded_file_path = fs.url(filename)

            # Process the CSV file
            with open(uploaded_file_path[1:], newline='') as csvfile:  # Remove leading '/' from URL
                reader = csv.reader(csvfile)
                for row in reader:
                    if len(row) >= 3:  # Ensure there are enough columns
                        customer_number = row[0]
                        name = row[1]
                        balance = row[2]
                        if balance.isdigit():  # Ensure balance is a valid number
                            Customer.objects.create(customer_number=customer_number, name=name, balance=int(balance))

            return redirect('customer_list')

    return render(request, 'store/add_customer.html')

from .forms import UploadCSVForm

def add_customers(request):
    if request.method == 'POST':
        form = UploadCSVForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            if not csv_file.name.endswith('.csv'):
                messages.error(request, 'This is not a CSV file.')
                return redirect('add_customers')

            # Read the CSV file
            decoded_file = csv_file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            next(io_string)  # Skip the header row if you have one

            for row in csv.reader(io_string, delimiter=','):
                # Assuming the CSV has customer_number, name, and balance
                customer_number, name, balance = row
                Customer.objects.create(
                    customer_number=customer_number,
                    name=name,
                    balance=float(balance)  # Convert balance to float
                )
            messages.success(request, 'Customers imported successfully!')
            return redirect('customer_list')  # Redirect to customer list or another page
    else:
        form = UploadCSVForm()
    
    return render(request, 'store/add_customers.html', {'form': form})

def edit_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    if request.method == "POST":
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()  # Update the item in the database
            return redirect('customer_list')  # Redirect to the item list page
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'store/edit_customer.html', {'customer': customer})
    
    # if request.method == "POST":
    #     customer_number = request.POST.get('customer_number')
    #     name = request.POST.get('name')
    #     balance = request.POST.get('balance')

    #     if name and balance.isdigit() and customer_number:
    #         customer.customer_number = customer_number
    #         customer.name = name
    #         customer.balance = int(balance)
    #         customer.save()  # Save the updated customer details
    #         return redirect('customer_list')

    # return render(request, 'store/edit_customer.html', {'customer': customer})


# def check_wallet_balance(request):
#     purchases = []
#     customer = None

#     if request.method == "POST":
#         search_query = request.POST.get('search_query')
#         try:
#             customer = Customer.objects.get(customer_number=search_query)  # Search by customer number
#         except Customer.DoesNotExist:
#             customer = Customer.objects.filter(name__icontains=search_query).first()  # Search by name

#         if customer:
#             purchases = ItemPurchase.objects.filter(customer=customer)  # Fetch purchases for the customer

#     return render(request, 'store/check_wallet_balance.html', {'customer': customer, 'purchases': purchases})

def delete_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    if request.method == "POST":
        customer.delete()  # Delete the customer
        return redirect('customer_list')
    return render(request, 'store/delete_customer.html', {'customer': customer})

def export_customers(request):
    # Create the HttpResponse object with the appropriate CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="customers.csv"'

    writer = csv.writer(response)
    writer.writerow(['Customer Number', 'Customer Name', 'Balance'])  # Write the header

    # Fetch all customers and write to the CSV
    customers = Customer.objects.all()
    for customer in customers:
        writer.writerow([customer.customer_number, customer.name, customer.balance])

    return response

def customer_list(request):
    customers = Customer.objects.all()
    return render(request, 'store/customer_list.html', {'customers': customers})

def add_money(request, customer_id):  # Accept customer_number as a parameter
    if request.method == 'POST':
        amount = request.POST.get('amount')

        # Fetch the customer
        customer = get_object_or_404(Customer, id=customer_id)

        try:
            # Convert amount to Decimal
            amount = Decimal(amount)
            if amount <= 0:
                messages.error(request, "Amount must be greater than zero.")
                return redirect('add_money', id=customer_id)

            # Update customer balance
            customer.balance += amount
            customer.save()

            # Create a transaction for adding money
            Transaction.objects.create(
                customer=customer,
                transaction_type='add',
                amount=amount
            )

            messages.success(request, "Money added successfully!")
            return redirect('check_wallet_balance')

        except (ValueError):
            messages.error(request, "Invalid amount.")
            return redirect('add_money', id=customer_id)

    return render(request, 'store/add_money.html')

def item_purchase(request):
    if request.method == 'POST':
        customer_number = request.POST.get('customer_number')
        item_ids = request.POST.getlist('item')  # Get list of item IDs
        quantities = request.POST.getlist('quantity')  # Get list of quantities

        # Fetch the customer
        customer = get_object_or_404(Customer, customer_number=customer_number)

        total_cost = 0
        items_to_purchase = []

        for item_id, quantity in zip(item_ids, quantities):
            if item_id:  # Ensure the item ID is not empty
                item = get_object_or_404(Item, id=item_id)
                total_cost += item.price * int(quantity)
                items_to_purchase.append((item, int(quantity)))

        # Check if the customer has sufficient balance
        if customer.balance < total_cost:
            messages.error(request, "Insufficient balance for this purchase.")
            return redirect('item_purchase')

        # Process the purchase and create transactions
        for item, quantity in items_to_purchase:
            Transaction.objects.create(
                customer=customer,
                transaction_type='purchase',
                amount=item.price * quantity,
                item=item
            )
            # Deduct the amount from the customer's balance
            customer.balance -= item.price * quantity
        
        # Save the updated customer balance
        customer.save()
        messages.success(request, "Purchase completed successfully!")
        return redirect('item_purchase')  # Redirect to a relevant page

    # For GET request, fetch all items
    items = Item.objects.all()
    return render(request, 'store/item_purchase.html', {'items': items})

def get_customer_details(request, customer_number):
    try:
        customer = Customer.objects.get(customer_number=customer_number)
        data = {
            'customer_name': customer.name,
            'balance': customer.balance,
        }
        return JsonResponse(data)
    except Customer.DoesNotExist:
        return JsonResponse({'customer_name': None, 'balance': None})

def check_wallet_balance(request):
    customer = None
    transactions = None
    if request.method == 'POST':
        customer_number = request.POST.get('customer_number')
        try:
            customer = Customer.objects.get(customer_number=customer_number)
            transactions = Transaction.objects.filter(customer=customer).order_by('-timestamp')
        except Customer.DoesNotExist:
            messages.error(request, "Customer not found.")
            return redirect('check_wallet_balance') # Redirect back to the form

    #Handle GET requests for exports
    elif request.GET.get('export'):
        customer_number = request.GET.get('customer_number')
        if not customer_number:
            return HttpResponseBadRequest("Customer number is required for export.")
        try:
            customer = Customer.objects.get(customer_number=customer_number)
            transactions = Transaction.objects.filter(customer=customer).order_by('-timestamp')
        except Customer.DoesNotExist:
            return HttpResponseBadRequest("Customer not found.")

    if request.GET.get('export') == 'excel':
        if not transactions:
            return HttpResponseBadRequest("No transactions to export.")
        return export_to_excel(transactions, customer)
    elif request.GET.get('export') == 'pdf':
        if not transactions:
            return HttpResponseBadRequest("No transactions to export.")
        return export_to_pdf(transactions)

    return render(request, 'store/check_wallet_balance.html', {'customer': customer, 'transactions': transactions})


def export_to_excel(transactions, customer):
    if not transactions:
        return HttpResponse("No transactions to export.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transaction History"

    # Styling elements
    header_style = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow header
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_left = Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Light blue
    light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid") # Light green

    # Customer Information Header
    sheet.merge_cells('A1:D1')  # Merge cells for header
    header_cell = sheet['A1']
    header_cell.value = f"Customer Information: {customer.customer_number} - {customer.name}"
    header_cell.font = Font(bold=True, size=14)
    header_cell.fill = light_blue_fill
    header_cell.alignment = alignment_center

    # Customer Details
    sheet['A2'] = "Customer Number:"
    sheet['B2'] = customer.customer_number
    sheet['C2'] = "Name:"
    sheet['D2'] = customer.name
    sheet['A3'] = "Balance:"
    sheet['B3'] = customer.balance

    for cell in ['A2', 'B2', 'C2', 'D2', 'A3', 'B3']:
        sheet[cell].alignment = alignment_left
        sheet[cell].font = Font(bold=True)
        sheet[cell].border = thin_border

    # Transaction Header Row
    sheet.append(["Type", "Item", "Amount", "Date"])
    for cell in sheet[4]:
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border

    # Transaction Data Rows
    for transaction in transactions:
        row = [
            transaction.get_transaction_type_display(),
            transaction.item.name if transaction.item else "N/A",
            transaction.amount,
            transaction.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
        ]
        sheet.append(row)

        # Access the last row added (which is the current transaction row)
        current_row = sheet.max_row
        for col in range(1, 5):  # Columns A to D
            cell = sheet.cell(row=current_row, column=col)
            cell.alignment = alignment_left
            cell.border = thin_border

    # Totals Calculation
    add_total = sum(t.amount for t in transactions if t.transaction_type == 'add')
    purchase_total = sum(t.amount for t in transactions if t.transaction_type == 'purchase')

    # Totals Row
    sheet.append(["Total Add Money:", "", add_total, ""])
    sheet.append(["Total Purchase:", "", purchase_total, ""])

    # Style the totals
    for row in range(sheet.max_row - 1, sheet.max_row + 1):  # Last two rows
        for col in range(1, 5):  # Columns A to D
            cell = sheet.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.alignment = alignment_center
            cell.border = thin_border
            if row == sheet.max_row - 1:
                cell.fill = light_green_fill
            else:
                cell.fill = light_blue_fill

    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = "transactions.xlsx"
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

def export_to_pdf(transactions):
    if not transactions:
        return HttpResponse("No transactions to export.")

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    c.drawString(250, 750, "Transaction History")
    c.setFont("Helvetica", 12)

    data = [["Type", "Item", "Amount", "Date"]]
    for transaction in transactions:
        data.append([
            transaction.get_transaction_type_display(),
            transaction.item.name if transaction.item else "N/A",
            transaction.amount,
            transaction.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    w, h = table.wrapOn(c, 10, 10)
    table.drawOn(c, 150, 700 - h)

    c.showPage()
    c.save()
    buffer.seek(0)
    return HttpResponse(buffer.getvalue(), content_type='application/pdf')

def get_customer_details(request, customer_number):
    try:
        customer = Customer.objects.get(customer_number=customer_number)
        data = {
            'customer_name': customer.name,
            'balance': customer.balance,
        }
        return JsonResponse(data)
    except Customer.DoesNotExist:
        return JsonResponse({'customer_name': None, 'balance': None})